""".xlsx — openpyxl 行列 → 自然语言描述
.csv  — chardet + csv.Sniffer 分隔符自适应
"""

import csv
import os
from typing import List

from .base import ContentElement, DocumentLoader


class XlsxLoader(DocumentLoader):
    """.xlsx 加载器：首行作列头，后续每行拼接为 "列头: 值" 格式。"""

    @property
    def supported_extensions(self) -> List[str]:
        return [".xlsx"]

    def load(self, file_path: str) -> List[ContentElement]:
        try:
            import openpyxl
        except ImportError:
            raise ImportError("请安装 openpyxl: pip install openpyxl")

        wb = openpyxl.load_workbook(file_path, data_only=True)
        elements: List[ContentElement] = []
        file_name = os.path.basename(file_path)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            headers = [str(h).strip() if h is not None else "" for h in rows[0]]

            for row_idx, row in enumerate(rows[1:], start=2):
                parts = []
                for col_idx, (header, value) in enumerate(zip(headers, row)):
                    if value is None or str(value).strip() == "":
                        continue
                    h = header if header else f"列{col_idx + 1}"
                    parts.append(f"{h}: {str(value).strip()}")
                if parts:
                    elements.append(ContentElement(
                        content=" | ".join(parts),
                        doc_type="table_row",
                        metadata={
                            "file_path": file_path,
                            "file_name": file_name,
                            "sheet": sheet_name,
                            "row": row_idx,
                        },
                    ))

        wb.close()
        return elements


class CsvLoader(DocumentLoader):
    """.csv 加载器：chardet 编码检测 + csv.Sniffer 分隔符自适应。"""

    @property
    def supported_extensions(self) -> List[str]:
        return [".csv"]

    def load(self, file_path: str) -> List[ContentElement]:
        try:
            import chardet
            with open(file_path, "rb") as f:
                raw = f.read()
            encoding = chardet.detect(raw)["encoding"] or "utf-8"
            content = raw.decode(encoding, errors="replace")
        except ImportError:
            with open(file_path, "r", encoding="utf-8", errors="replace") as f:
                content = f.read()

        # 分隔符检测
        sniffer = csv.Sniffer()
        dialect = sniffer.sniff(content[:8192])

        lines = content.strip().split("\n")
        if not lines:
            return []

        reader = csv.reader(lines, delimiter=dialect.delimiter)
        rows = list(reader)
        if not rows:
            return []

        headers = [h.strip() for h in rows[0]]
        elements: List[ContentElement] = []
        file_name = os.path.basename(file_path)

        for row_idx, row in enumerate(rows[1:], start=2):
            parts = []
            for col_idx, (header, value) in enumerate(zip(headers, row)):
                if not value or not value.strip():
                    continue
                h = header if header else f"列{col_idx + 1}"
                parts.append(f"{h}: {value.strip()}")
            if parts:
                elements.append(ContentElement(
                    content=" | ".join(parts),
                    doc_type="table_row",
                    metadata={
                        "file_path": file_path,
                        "file_name": file_name,
                        "row": row_idx,
                    },
                ))

        return elements
